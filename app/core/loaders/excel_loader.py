"""Pandas Excel reader.

Pandas parser for .xlsx files.

"""
from pathlib import Path
from typing import Any, List, Optional, Union, BinaryIO, IO

from llama_index.core.readers.base import BaseReader

from core.base import Document
from datetime import datetime


class PandasExcelReader(BaseReader):
    r"""Pandas-based CSV parser.

    Parses CSVs using the separator detection from Pandas `read_csv` function.
    If special parameters are required, use the `pandas_config` dict.

    Args:

        pandas_config (dict): Options for the `pandas.read_excel` function call.
            Refer to https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html
            for more information. Set to empty dict by default,
            this means defaults will be used.

    """

    def __init__(
        self,
        *args: Any,
        pandas_config: Optional[dict] = None,
        row_joiner: str = "\n",
        col_joiner: str = " ",
        **kwargs: Any,
    ) -> None:
        """Init params."""
        super().__init__(*args, **kwargs)
        self._pandas_config = pandas_config or {}
        self._row_joiner = row_joiner if row_joiner else "\n"
        self._col_joiner = col_joiner if col_joiner else " "

    def load_data(
        self,
        file: Union[Path, str, BinaryIO, IO],
        include_sheetname: bool = False,
        sheet_name: Optional[Union[str, int, list]] = None,
        extra_info: Optional[dict] = None,
        **kwargs,
    ) -> List[Document]:
        """Parse file and extract values from a specific column.

        Args:
            file (Path): The path to the Excel file to read.
            include_sheetname (bool): Whether to include the sheet name in the output.
            sheet_name (Union[str, int, None]): The specific sheet to read from,
                default is None which reads all sheets.

        Returns:
            List[Document]: A list of`Document objects containing the
                values from the specified column in the Excel file.
        """
        import itertools

        try:
            import pandas as pd
        except ImportError:
            raise ImportError(
                "install pandas using `pip3 install pandas` to use this loader"
            )

        if sheet_name is not None:
            sheet_name = (
                [sheet_name] if not isinstance(sheet_name, list) else sheet_name
            )

        if isinstance(file, (str, Path)):
            file = Path(file)

        dfs = pd.read_excel(file, sheet_name=sheet_name, **self._pandas_config)
        sheet_names = dfs.keys()
        df_sheets = []

        for key in sheet_names:
            sheet = []
            if include_sheetname:
                sheet.append([key])
            dfs[key] = dfs[key].dropna(axis=0, how="all")
            dfs[key].fillna("", inplace=True)
            sheet.extend(dfs[key].values.astype(str).tolist())
            df_sheets.append(sheet)

        text_list = list(
            itertools.chain.from_iterable(df_sheets)
        )  # flatten list of lists

        output = [
            Document(
                text=self._row_joiner.join(
                    self._col_joiner.join(sublist) for sublist in text_list
                ),
                metadata=extra_info or {},
            )
        ]

        return output

def lower_preserve_urls(text: str) -> str:
        """
        This function converts all text to lowercase but keeps URLs intact.
        URLs are identified by information patterns that start with http or https.
        """
        import re
        url_pattern = r'(https?://[^\s]+)'
        matches = list(re.finditer(url_pattern, text))
        if not matches:
            return text.lower()
        new_text = []
        last_end = 0
        for match in matches:
            new_text.append(text[last_end:match.start()].lower())
            new_text.append(match.group())
            last_end = match.end()

        new_text.append(text[last_end:].lower())
        return "".join(new_text)

class ExcelReader(BaseReader):
    r"""Spreadsheet exporter respecting multiple worksheets

    Parses CSVs using the separator detection from Pandas `read_csv` function.
    If special parameters are required, use the `pandas_config` dict.

    Args:

        pandas_config (dict): Options for the `pandas.read_excel` function call.
            Refer to https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html
            for more information. Set to empty dict by default,
            this means defaults will be used.

    """

    def __init__(
        self,
        *args: Any,
        pandas_config: Optional[dict] = None,
        row_joiner: str = "\n",
        col_joiner: str = " ",
        **kwargs: Any,
    ) -> None:
        """Init params."""
        super().__init__(*args, **kwargs)
        self._pandas_config = pandas_config or {}
        self._row_joiner = row_joiner if row_joiner else "\n"
        self._col_joiner = col_joiner if col_joiner else " "

    def load_data(
        self,
        file: Union[Path, str, BinaryIO, IO],
        include_sheetname: bool = True,
        sheet_name: Optional[Union[str, int, list]] = None,
        extra_info: Optional[dict] = None,
        **kwargs,
    ) -> List[Document]:
        """Parse file and extract values from a specific column.

        Args:
            file (Path): The path to the Excel file to read.
            include_sheetname (bool): Whether to include the sheet name in the output.
            sheet_name (Union[str, int, None]): The specific sheet to read from,
                default is None which reads all sheets.

        Returns:
            List[Document]: A list of`Document objects containing the
                values from the specified column in the Excel file.
        """

        try:
            import pandas as pd
            import openpyxl
        except ImportError:
            raise ImportError(
                "install pandas and openpyxl using `pip install pandas openpyxl` to use this loader"
            )

        if sheet_name is not None:
            sheet_name = (
                [sheet_name] if not isinstance(sheet_name, list) else sheet_name
            )

        if isinstance(file, (str, Path)):
            file = Path(file)
            file_name = file.name
        else:
            file_name = extra_info.get("file_name", "unknown")

        extra_info = extra_info or {}

        try:
            if not isinstance(file, (str, Path)):
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(file.read() if hasattr(file, 'read') else file)
                    tmp_path = tmp.name
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                import os
                os.unlink(tmp_path)
            else:
                wb = openpyxl.load_workbook(file, data_only=True)

            if sheet_name is None:
                sheet_names_to_process = wb.sheetnames
            else:
                sheet_names_to_process = []
                for name in sheet_name:
                    if isinstance(name, int) and 0 <= name < len(wb.sheetnames):
                        sheet_names_to_process.append(wb.sheetnames[name])
                    elif name in wb.sheetnames:
                        sheet_names_to_process.append(name)
        except Exception as e:
            dfs = pd.read_excel(file, sheet_name=sheet_name, **self._pandas_config)
            return self._process_dataframes(dfs, file_name, include_sheetname, extra_info)

        output = []
        for sheet_name in sheet_names_to_process:
            sheet = wb[sheet_name]

            merged_cells = {}
            for merged_range in sheet.merged_cells.ranges:
                top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                if top_left_cell.value is not None:
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            merged_cells[(row, col)] = top_left_cell.value

            max_row = sheet.max_row
            max_col = sheet.max_column

            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(1, col).value
                if cell_value is None and (1, col) in merged_cells:
                    cell_value = merged_cells[(1, col)]
                headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")

            data_rows = []
            for row_idx in range(2, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell_value = sheet.cell(row_idx, col_idx).value

                    if cell_value is None and (row_idx, col_idx) in merged_cells:
                        cell_value = merged_cells[(row_idx, col_idx)]

                    row_data.append(str(cell_value) if cell_value is not None else "")

                if any(cell != "" for cell in row_data):
                    data_rows.append(row_data)

            for idx, row in enumerate(data_rows):
                content = ''
                for cell_idx, cell_text in enumerate(row):
                    if cell_text:
                        content += f"""{headers[cell_idx]}:{self._row_joiner}{cell_text}{self._row_joiner}{self._row_joiner}"""

                if include_sheetname:
                    content = f"Sheet {sheet_name} of file {file_name}\n{content}"

                metadata = {
                    "sheet_name": sheet_name,
                    "idx": idx,
                    "uploaded_at": datetime.now(),
                    **extra_info
                }

                if content.strip():
                    output.append(Document(text=lower_preserve_urls(content), metadata=metadata))

        if not output:
            dfs = pd.read_excel(file, sheet_name=sheet_name, **self._pandas_config)
            return self._process_dataframes(dfs, file_name, include_sheetname, extra_info)

        return output

    def _process_dataframes(self, dfs, file_name, include_sheetname, extra_info):
        """Xử lý các DataFrames từ pandas"""
        output = []

        for key in dfs.keys():
            dfs[key] = dfs[key].dropna(axis=0, how='all')
            dfs[key] = dfs[key].astype("object")

            dfs[key] = dfs[key].fillna(method='ffill', axis=0)
            dfs[key] = dfs[key].fillna(method='ffill', axis=1)
            dfs[key].fillna("", inplace=True)

            rows = dfs[key].values.astype(str).tolist()
            columns = dfs[key].columns.tolist()

            for idx, row in enumerate(rows):
                content = ''
                for cell_idx, cell_text in enumerate(row):
                    if cell_text:
                        content += f"{columns[cell_idx]}:{self._row_joiner}{cell_text}{self._row_joiner}{self._row_joiner}"

                if include_sheetname:
                    content = f"Sheet {key} của file {file_name}\n\n{content}"

                metadata = {
                    "sheet_name": key,
                    "idx": idx,
                    "uploaded_at": datetime.now(),
                    **extra_info
                }

                if content.strip():
                    output.append(Document(text=lower_preserve_urls(content), metadata=metadata))

        return output
